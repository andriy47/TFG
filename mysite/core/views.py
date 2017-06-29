import facebook
import json
import requests
import openpyxl
import operator
import datetime
import collections

from django.contrib.auth.decorators import login_required
# from django.contrib.auth.forms import AdminPasswordChangeForm, PasswordChangeForm, UserCreationForm
from django.contrib.auth import update_session_auth_hash, login, authenticate
from django.contrib import messages
from django.shortcuts import render, redirect
from social_django.models import UserSocialAuth
from pymongo import MongoClient
from django.shortcuts import render_to_response

from mysite.form import InputNumeroForm

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = '/home/and/TFG/mysite/templates/registration/emotions.xlsx', read_only=True)

sheets = wb.sheetnames
ws = wb[sheets[0]]
diccionary = dict()

#Read de xlsx file and save in arrayhash the words and her puntuation in the table.
for row in ws.iter_rows(min_row=2, max_col=11, max_row=14183):
    diccionary[row[0].value] = []
    for cell in row:
        if not cell.value in diccionary:
            diccionary[row[0].value].append(cell.value)


client = MongoClient('localhost', 27017)
db = client['analistics']
access_token = 'EAACEdEose0cBAGZAQ79tTLAWQxKQvDg3DLmowg4i2kZCxity03TgoG9NkuOXbBO8csZBgZAZA151CrioV8Gcjqz2P6qDvPORj8ACsS50s1HrOMv3oZCKv8nInJ0JtEslXxx8yH5IaNQWpeoPhZBaE9wu0f735XMRuJsXKe4ir4cHjZBQC8JZBPUpYmVCaDxnNRicZD'
@login_required
def home(request):
    return render(request, 'core/home.html')

@login_required
def datag(request):
    if request.method == 'POST':
        if request.POST.get('numero', False):
            id_group = int(request.POST['numero'])
    graph = facebook.GraphAPI(access_token)
    mesage = dict()
    finaldata = dict()
    usersPart = dict()
    choice = False
    result = db.resultsgroups.find_one({'id_group':id_group})
    if result:
        # print(result)
        if result['date'] == datetime.datetime.now().date().isoformat():
             choice = True
    if not choice:
        #Start to GET all comments from posts with diferents levels of Facebook groups.
        datass = graph.get_connections(id_group, '?fields=feed{comments{message,from,created_time}}')
        datass = datass.get('feed')

        while(True):
        	try:
        		for dat in datass.get('data'):
        			datt = dat.get('comments')
        			if datt:
        				for allcoments in datt.get('data'):
        					id_user = allcoments.get('from').get('id')
	    		                        created_date = allcoments['created_time']
        					user_message = allcoments.get('message').split()
        					if id_user in mesage:
        						d = mesage.get(id_user)[0] + user_message
        						mesage[id_user] = [list(set(d)), created_date]
        					else:
        						mesage[id_user] =  [user_message, created_date]
        		if datass:
        			datass = requests.get(datass['paging']['next']).json()
        	except KeyError:
        		break
        #Start to GET all reply comments from posts of Facebook groups.
        comments = graph.get_connections(id_group, '?fields=feed{comments{comments}}')
        comments = comments.get('feed')
        while(True):
        	try:
        		for val in comments.get('data'):
        			if val.get('comments'):
        				if val.get('comments').get('data')[0].get('comments'):
        					for replys in val.get('comments').get('data')[0].get('comments').get('data'):
        						id_com_usr = replys.get('from').get('id')
        	            				reated_date = replys.get('created_time')
        						mess_usr = replys.get('message').split()
        						if id_com_usr in mesage:
        							t = mesage.get(id_com_usr)[0] + mess_usr
        							mesage[id_com_usr] = [list(set(t)), created_date]
        						else:
        							mesage[id_com_usr] = [mess_usr, created_date]
        		if comments:
        			comments = requests.get(comments.get('paging')['next']).json()
        	except TypeError:
        		break
        #FINAL ARRAYHASH WITH ALL COMMENTS AT PERSONS
        sentimientos = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for key, valu in mesage.iteritems():
        	for pal in valu[0]:
        		if pal in diccionary:
	                    	db.datagroup.update_one(
	    	                {
            	        	        "id_user" : key,
                	    	        "id_group" : id_group,
        	                    	"date_created" : valu[1],
	    	                    	"content" : pal
	            	        },
	                 	{
	                            	"$set" : {"content" : pal}
    	                    	},
            	        	        upsert=True
                	    	)
       				sentimientos = map(operator.add, sentimientos, diccionary.get(pal))
        typeemoji = ['Positivo','Confiado','Alegria','Anticipacion','Sopresa','Tristesa','Disgusto','Miedo','Ira','Negatividad']
        allfound = sum(sentimientos)
        for typ, punt in zip(typeemoji,sentimientos):
            finaldata[typ] = round(punt * 100/float(allfound), 2)
        db.resultsgroups.update_one(
        {
            "id_group" : id_group
        },
        {
            "$set" : {"id_group" : id_group, "finaldata": finaldata, "date" : datetime.datetime.now().date().isoformat()}
         },
            upsert=True
        )

    else:
        resultdat = db.resultsgroups.find({'id_group': id_group})
        for res in resultdat:
            finaldata = res['finaldata']
    usersInGroup = db.datagroup.distinct("id_user");
    for var1 in usersInGroup:
        nameUser = graph.get_object(id=var1)
        adat = db.datagroup.count({"id_user": var1, "id_group": id_group})
        usersPart[adat] = nameUser['name']
    usersPart = collections.OrderedDict(reversed(sorted(usersPart.items())))

    return render(request, 'core/datag.html', {
    'ladate': finaldata,
    'usersPart' : usersPart
    })

@login_required
def settings(request):
    #access_token = '407156876319284|h5wEvmcAky9NjDDXXp4rnjHkFbg'
    user = request.user
    try:
        facebook_login = user.social_auth.get(provider='facebook')
    except UserSocialAuth.DoesNotExist:
        facebook_login = None

    #Get tokens and id necesry from facebook user

    graph = facebook.GraphAPI(access_token)
    user_id = '786814664672852'
    profile = graph.get_object(user_id)

    #Check if user are yet in DataBase
    result_user = db.users.find_one({'id':user_id})

    #Insert if not exist
    if str(result_user) == 'None':
	       db.users.insert(profile)
    groups = graph.get_connections(user_id, 'groups')
    can_disconnect = (user.social_auth.count() > 1 or user.has_usable_password())

    groups_avaliables = {}
    #Add user_id wich groups is had
    for value in groups['data']:
        result = db.groups.find_one({'id':value['id']})
        groups_avaliables[value['id']]=value['name']
        if str(result) == 'None':
            value['user_id'] = user
            db.groups.insert(value)
    return render(request, 'core/settings.html', {
        'groups_avaliables': groups_avaliables,
        'facebook_login': facebook_login,
        'can_disconnect': can_disconnect
        })
