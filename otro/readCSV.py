import openpyxl
import operator
import json

# wb = openpyxl.Workbook()
# wb = openpyxl.load_workbook(filename = 'emotions.xlsx', read_only=True)
#
# sheets = wb.sheetnames
# ws = wb[sheets[0]]
#
# data = dict()
#
# #Read de xlsx file and save in arrayhash the words and her puntuation in the table.
# for row in ws.iter_rows(min_row=2, max_col=11, max_row=14183):
#     data[row[0].value] = []
#     for cell in row:
#         if not cell.value in data:
#             data[row[0].value].append(cell.value)
#
# print(data.get('estragos'))
data = dict()
data1 = {'hola':['todo','bien','mal']}
data2 = ['todo', 'genial']


t= data1.get('hola') + data2
data['hola'] = list(set(t))
print(data)
