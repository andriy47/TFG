#Imports for xlsx File.
import openpyxl
import operator

#Imports for GET data from Facebook and parse json.
import facebook
import requests
import json
import pprint

#Requirements to connect at MongoClient.
from collections import OrderedDict
from pymongo import MongoClient

client = MongoClient('localhost', 27017)
db = client['analistics']

#Temporal tokens to use.
access_token = 'EAACEdEose0cBANRcTIIpnZCpsECP3abRYGV6H8JQFBRp95erYqNMhYvlELh07bySSKMv7hiRYZBCFimgU6xYf5Ed8e3TeauYZCwBaHFK55TNawZAL5ZBDEr9AVUZBsCj9opErv5t2IVZAgkb3Vx18HLnRCVXrHya0HxZBKDxWuM4lsk1fOJ9zh8lWPyFRbD9p3oZD'
#user = '1469643709723274'
id_group = '1038566552861343'

mesage = dict()
graph = facebook.GraphAPI(access_token)

#Start to GET all comments from posts with diferents levels of Facebook groups.
datass = graph.get_connections(id_group, '?fields=feed{comments{message,from,created_time}}')
datass = datass.get('feed')

while(True):
    try:
        for dat in datass.get('data'):
            datt = dat.get('comments')
            if datt:
                for allcoments in datt.get('data'):
                    id_user = allcoments.get('from').get('id')
                    created_date = allcoments['created_time']
                    user_message = allcoments.get('message').split()
                    if id_user in mesage:
                        d = mesage.get(id_user)[0] + user_message
                        mesage[id_user] = [list(set(d)), created_date]
                    else:
                        mesage[id_user] = [user_message, created_date]
                #     db.datagroup.update_one(
                #         {
				# "id_user" : id_user,
				# "id_group" : id_group,
				# "date_created" : created_date,
				# "content" : mesage[id_user]
                #         },
                #         {
				# "$set" : {"content" : mesage[id_user]}
                #         },
                #         upsert=True
                #     );
        if datass:
            datass = requests.get(datass['paging']['next']).json()
    except KeyError:
        break

#Start to GET all reply comments from posts of Facebook groups.
comments = graph.get_connections(id_group, '?fields=feed{comments{comments}}')
comments = comments.get('feed')
while(True):
    try:
        for val in comments.get('data'):
            if val.get('comments'):
                if val.get('comments').get('data')[0].get('comments'):
                    for replys in val.get('comments').get('data')[0].get('comments').get('data'):
                        id_com_usr = replys.get('from').get('id')
                        created_date = replys.get('created_time')
                        mess_usr = replys.get('message').split()
                        if id_com_usr in mesage:
                            t = mesage.get(id_com_usr)[0] + mess_usr
                            mesage[id_com_usr] = [list(set(t)), created_date]
                        else:
                            mesage[id_com_usr] = [mess_usr, created_date]
			# db.datagroup.update_one(
			# {
			# 	"id_user" : id_com_usr,
	        #                 "id_group" : id_group,
	        #                 "date_created" : created_date,
			# 	"content" : mesage[id_com_usr]
            #             },
            #             {
            #                 "$set" : {"content" : mesage[id_com_usr]}
            #             },
	        #                 upsert=True
	        #         )
        if comments:
            comments = requests.get(comments.get('paging')['next']).json()
    except TypeError:
        break
#FINAL ARRAYHASH WITH ALL COMMENTS AT PERSONS
#print(json.dumps(mesage))
print("1st Done!")

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = 'emotions.xlsx', read_only=True)

sheets = wb.sheetnames
ws = wb[sheets[0]]
diccionary = dict()

#Read de xlsx file and save in arrayhash the words and her puntuation in the table.
for row in ws.iter_rows(min_row=2, max_col=11, max_row=14183):
    diccionary[row[0].value] = []
    for cell in row:
        if not cell.value in diccionary:
            diccionary[row[0].value].append(cell.value)

a=0
print('2nd Done!')
sentimientos = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
for key, valu in mesage.items():
    for pal in valu[0]:
        if pal in diccionary:
            a=a+1
            sentimientos = map(operator.add, sentimientos, diccionary.get(pal))

print(a)
#print(sentimientos)
